from zipfile import BadZipFile

from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from rest_framework import serializers as drf_serializers

from users.api.v1.serializers import InvitationCreateSerializer
from users.choices import InvitationStatus
from users.constants import MAX_BULK_INVITE_ROWS
from users.models import Invitation
from users.tasks import send_invitation_email_task

User = get_user_model()


def parse_invitation_emails(file):
    """Read a single-column .xlsx upload into a list of raw email strings,
    skipping a header row if present. Raises `ValueError` if the file isn't
    a readable .xlsx workbook, or has more than `MAX_BULK_INVITE_ROWS` rows."""
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except (BadZipFile, KeyError, InvalidFileException) as exc:
        raise ValueError("file must be a valid .xlsx workbook.") from exc

    worksheet = workbook.active

    emails = [
        str(cell_value).strip()
        for (cell_value,) in worksheet.iter_rows(min_col=1, max_col=1, values_only=True)
        if cell_value is not None and str(cell_value).strip()
    ]

    if emails and "@" not in emails[0]:
        emails = emails[1:]

    if len(emails) > MAX_BULK_INVITE_ROWS:
        raise ValueError(f"file must contain at most {MAX_BULK_INVITE_ROWS} emails.")

    return emails


def bulk_create_invitations(emails, request):
    """Create a pending Invitation for each valid, non-duplicate email via
    the same path as single invitation creation. Invalid, duplicate,
    already-invited/existing-user emails, and rows rejected by
    `InvitationCreateSerializer` (e.g. the org's pending-invitation cap)
    are skipped and reported instead of failing the batch."""
    organization = request.user.organization
    created_invitations = []
    skipped_rows = []
    seen_emails = set()

    for original_email in emails:
        email = original_email.strip()
        normalized_email = email.lower()

        try:
            validate_email(email)
        except ValidationError:
            skipped_rows.append({"email": original_email, "reason": "invalid email"})
            continue

        if normalized_email in seen_emails:
            skipped_rows.append(
                {"email": original_email, "reason": "duplicate in file"}
            )
            continue
        seen_emails.add(normalized_email)

        if User.objects.filter(organization=organization, email__iexact=email).exists():
            skipped_rows.append(
                {
                    "email": original_email,
                    "reason": "user already exists in organization",
                }
            )
            continue

        if (
            Invitation.objects.for_organization(organization)
            .filter(email__iexact=email, status=InvitationStatus.PENDING)
            .exists()
        ):
            skipped_rows.append(
                {"email": original_email, "reason": "invitation already pending"}
            )
            continue

        serializer = InvitationCreateSerializer(
            data={"email": email}, context={"request": request}
        )
        try:
            serializer.is_valid(raise_exception=True)
        except drf_serializers.ValidationError as exc:
            field_errors = next(iter(exc.detail.values()))
            skipped_rows.append(
                {"email": original_email, "reason": str(field_errors[0])}
            )
            continue

        invitation = serializer.save()
        send_invitation_email_task.delay(invitation.pk)
        created_invitations.append(invitation)

    return {"created": created_invitations, "skipped": skipped_rows}
